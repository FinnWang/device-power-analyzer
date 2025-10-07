#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
無線滑鼠耗電分析 GUI 工具
整合所有分析功能的圖形化介面
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import seaborn as sns
from pathlib import Path
import threading
import queue
import os
import sys
from datetime import datetime

# 導入現有的分析模組
try:
    from mouse_power_analyzer import MousePowerAnalyzer
except ImportError:
    print("警告：無法導入 mouse_power_analyzer，將使用內建分析功能")
    MousePowerAnalyzer = None

class MousePowerGUI:
    """無線滑鼠耗電分析 GUI 主類別"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("無線滑鼠耗電分析工具 v1.0")
        self.root.geometry("1200x800")
        
        # 初始化變數
        self.analyzer = MousePowerAnalyzer() if MousePowerAnalyzer else None
        self.loaded_files = {}
        self.analysis_results = {}
        self.output_dir = "./gui_analysis_results"
        
        # 建立輸出目錄
        Path(self.output_dir).mkdir(exist_ok=True)
        
        # 設定樣式
        self.setup_styles()
        
        # 建立GUI元件
        self.create_widgets()
        
        # 初始化狀態
        self.update_status("就緒 - 請載入CSV檔案開始分析")
    
    def setup_styles(self):
        """設定GUI樣式"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # 自訂樣式
        style.configure('Title.TLabel', font=('Arial', 14, 'bold'))
        style.configure('Header.TLabel', font=('Arial', 12, 'bold'))
        style.configure('Success.TLabel', foreground='green')
        style.configure('Error.TLabel', foreground='red')
    
    def create_widgets(self):
        """建立GUI元件"""
        
        # 主標題
        title_frame = ttk.Frame(self.root)
        title_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(title_frame, text="無線滑鼠耗電分析工具", 
                 style='Title.TLabel').pack()
        
        # 建立筆記本（分頁）
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=5)
        
        # 分頁1: 檔案載入與基本資訊
        self.create_file_tab()
        
        # 分頁2: 單檔分析
        self.create_single_analysis_tab()
        
        # 分頁3: 多檔比較
        self.create_comparison_tab()
        
        # 分頁4: 報告生成
        self.create_report_tab()
        
        # 狀態列
        self.create_status_bar()
    
    def create_file_tab(self):
        """建立檔案載入分頁"""
        
        file_frame = ttk.Frame(self.notebook)
        self.notebook.add(file_frame, text="檔案管理")
        
        # 檔案載入區域
        load_frame = ttk.LabelFrame(file_frame, text="檔案載入", padding=10)
        load_frame.pack(fill='x', padx=10, pady=5)
        
        # 按鈕區域
        btn_frame = ttk.Frame(load_frame)
        btn_frame.pack(fill='x', pady=5)
        
        ttk.Button(btn_frame, text="載入單一檔案", 
                  command=self.load_single_file).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="載入多個檔案", 
                  command=self.load_multiple_files).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="載入資料夾", 
                  command=self.load_folder).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="清除所有檔案", 
                  command=self.clear_files).pack(side='left', padx=5)
        
        # 已載入檔案列表
        list_frame = ttk.LabelFrame(file_frame, text="已載入檔案", padding=10)
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        # 建立樹狀檢視
        columns = ('檔案名稱', '模式', '資料點數', '時間範圍', '平均功率')
        self.file_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=8)
        
        for col in columns:
            self.file_tree.heading(col, text=col)
            self.file_tree.column(col, width=120)
        
        # 滾動條
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.file_tree.yview)
        self.file_tree.configure(yscrollcommand=scrollbar.set)
        
        self.file_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # 檔案操作按鈕
        file_btn_frame = ttk.Frame(list_frame)
        file_btn_frame.pack(fill='x', pady=5)
        
        ttk.Button(file_btn_frame, text="檢視資料", 
                  command=self.view_data).pack(side='left', padx=5)
        ttk.Button(file_btn_frame, text="移除選中", 
                  command=self.remove_selected).pack(side='left', padx=5)
    
    def create_single_analysis_tab(self):
        """建立單檔分析分頁"""
        
        single_frame = ttk.Frame(self.notebook)
        self.notebook.add(single_frame, text="單檔分析")
        
        # 檔案選擇
        select_frame = ttk.LabelFrame(single_frame, text="選擇分析檔案", padding=10)
        select_frame.pack(fill='x', padx=10, pady=5)
        
        self.single_file_var = tk.StringVar()
        self.single_file_combo = ttk.Combobox(select_frame, textvariable=self.single_file_var, 
                                             state='readonly', width=50)
        self.single_file_combo.pack(side='left', padx=5)
        
        ttk.Button(select_frame, text="開始分析", 
                  command=self.run_single_analysis).pack(side='left', padx=5)
        
        # 分析選項
        options_frame = ttk.LabelFrame(single_frame, text="分析選項", padding=10)
        options_frame.pack(fill='x', padx=10, pady=5)
        
        self.include_detailed = tk.BooleanVar(value=True)
        self.include_statistics = tk.BooleanVar(value=True)
        self.include_battery = tk.BooleanVar(value=True)
        
        ttk.Checkbutton(options_frame, text="詳細分析圖表", 
                       variable=self.include_detailed).pack(side='left', padx=10)
        ttk.Checkbutton(options_frame, text="統計摘要", 
                       variable=self.include_statistics).pack(side='left', padx=10)
        ttk.Checkbutton(options_frame, text="電池續航估算", 
                       variable=self.include_battery).pack(side='left', padx=10)
        
        # 結果顯示區域
        result_frame = ttk.LabelFrame(single_frame, text="分析結果", padding=10)
        result_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.single_result_text = scrolledtext.ScrolledText(result_frame, height=15)
        self.single_result_text.pack(fill='both', expand=True)
    
    def create_comparison_tab(self):
        """建立多檔比較分頁"""
        
        comp_frame = ttk.Frame(self.notebook)
        self.notebook.add(comp_frame, text="多檔比較")
        
        # 檔案選擇
        select_frame = ttk.LabelFrame(comp_frame, text="選擇比較檔案", padding=10)
        select_frame.pack(fill='x', padx=10, pady=5)
        
        # 可選檔案列表
        self.comparison_listbox = tk.Listbox(select_frame, selectmode='multiple', height=6)
        self.comparison_listbox.pack(fill='x', pady=5)
        
        btn_frame = ttk.Frame(select_frame)
        btn_frame.pack(fill='x', pady=5)
        
        ttk.Button(btn_frame, text="全選", 
                  command=self.select_all_files).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="清除選擇", 
                  command=self.clear_selection).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="開始比較分析", 
                  command=self.run_comparison_analysis).pack(side='left', padx=5)
        
        # 比較選項
        comp_options_frame = ttk.LabelFrame(comp_frame, text="比較選項", padding=10)
        comp_options_frame.pack(fill='x', padx=10, pady=5)
        
        self.include_power_comparison = tk.BooleanVar(value=True)
        self.include_efficiency = tk.BooleanVar(value=True)
        self.include_battery_comparison = tk.BooleanVar(value=True)
        
        ttk.Checkbutton(comp_options_frame, text="功率比較", 
                       variable=self.include_power_comparison).pack(side='left', padx=10)
        ttk.Checkbutton(comp_options_frame, text="效率分析", 
                       variable=self.include_efficiency).pack(side='left', padx=10)
        ttk.Checkbutton(comp_options_frame, text="續航比較", 
                       variable=self.include_battery_comparison).pack(side='left', padx=10)
        
        # 結果顯示
        comp_result_frame = ttk.LabelFrame(comp_frame, text="比較結果", padding=10)
        comp_result_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.comp_result_text = scrolledtext.ScrolledText(comp_result_frame, height=15)
        self.comp_result_text.pack(fill='both', expand=True)
    
    def create_report_tab(self):
        """建立報告生成分頁"""
        
        report_frame = ttk.Frame(self.notebook)
        self.notebook.add(report_frame, text="報告生成")
        
        # 報告設定
        settings_frame = ttk.LabelFrame(report_frame, text="報告設定", padding=10)
        settings_frame.pack(fill='x', padx=10, pady=5)
        
        # 輸出目錄選擇
        dir_frame = ttk.Frame(settings_frame)
        dir_frame.pack(fill='x', pady=5)
        
        ttk.Label(dir_frame, text="輸出目錄:").pack(side='left')
        self.output_dir_var = tk.StringVar(value=self.output_dir)
        ttk.Entry(dir_frame, textvariable=self.output_dir_var, width=40).pack(side='left', padx=5)
        ttk.Button(dir_frame, text="瀏覽", 
                  command=self.select_output_dir).pack(side='left', padx=5)
        
        # 報告格式選項
        format_frame = ttk.Frame(settings_frame)
        format_frame.pack(fill='x', pady=5)
        
        ttk.Label(format_frame, text="報告格式:").pack(side='left')
        
        self.export_png = tk.BooleanVar(value=True)
        self.export_pdf = tk.BooleanVar(value=False)
        self.export_excel = tk.BooleanVar(value=True)
        
        ttk.Checkbutton(format_frame, text="PNG圖片", 
                       variable=self.export_png).pack(side='left', padx=10)
        ttk.Checkbutton(format_frame, text="PDF報告", 
                       variable=self.export_pdf).pack(side='left', padx=10)
        ttk.Checkbutton(format_frame, text="Excel統計", 
                       variable=self.export_excel).pack(side='left', padx=10)
        
        # 生成按鈕
        generate_frame = ttk.Frame(settings_frame)
        generate_frame.pack(fill='x', pady=10)
        
        ttk.Button(generate_frame, text="生成完整報告", 
                  command=self.generate_full_report).pack(side='left', padx=5)
        ttk.Button(generate_frame, text="開啟輸出目錄", 
                  command=self.open_output_dir).pack(side='left', padx=5)
        
        # 報告預覽
        preview_frame = ttk.LabelFrame(report_frame, text="報告預覽", padding=10)
        preview_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.report_text = scrolledtext.ScrolledText(preview_frame, height=15)
        self.report_text.pack(fill='both', expand=True)
    
    def create_status_bar(self):
        """建立狀態列"""
        
        self.status_frame = ttk.Frame(self.root)
        self.status_frame.pack(fill='x', side='bottom')
        
        self.status_var = tk.StringVar()
        self.status_label = ttk.Label(self.status_frame, textvariable=self.status_var)
        self.status_label.pack(side='left', padx=10, pady=5)
        
        # 進度條
        self.progress = ttk.Progressbar(self.status_frame, mode='indeterminate')
        self.progress.pack(side='right', padx=10, pady=5)
    
    def update_status(self, message):
        """更新狀態列"""
        self.status_var.set(f"狀態: {message}")
        self.root.update_idletasks()
    
    def start_progress(self):
        """開始進度條動畫"""
        self.progress.start()
    
    def stop_progress(self):
        """停止進度條動畫"""
        self.progress.stop()
    
    def load_single_file(self):
        """載入單一檔案"""
        
        file_path = filedialog.askopenfilename(
            title="選擇CSV檔案",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if file_path:
            self.load_file(file_path)
    
    def load_multiple_files(self):
        """載入多個檔案"""
        
        file_paths = filedialog.askopenfilenames(
            title="選擇多個CSV檔案",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        for file_path in file_paths:
            self.load_file(file_path)
    
    def load_folder(self):
        """載入資料夾中的所有CSV檔案"""
        
        folder_path = filedialog.askdirectory(title="選擇包含CSV檔案的資料夾")
        
        if folder_path:
            csv_files = list(Path(folder_path).glob("*.csv"))
            
            if csv_files:
                for file_path in csv_files:
                    self.load_file(str(file_path))
                
                messagebox.showinfo("載入完成", f"成功載入 {len(csv_files)} 個CSV檔案")
            else:
                messagebox.showwarning("警告", "選擇的資料夾中沒有找到CSV檔案")
    
    def load_file(self, file_path):
        """載入單一檔案的核心函數"""
        
        try:
            self.update_status(f"載入檔案: {Path(file_path).name}")
            
            # 讀取CSV檔案
            df = pd.read_csv(file_path)
            
            # 檢查欄位
            if len(df.columns) >= 4:
                df.columns = ['Time', 'Voltage', 'Current', 'Power']
            else:
                raise ValueError("CSV檔案格式不正確，需要至少4個欄位")
            
            # 推測模式名稱
            filename = Path(file_path).stem
            mode_name = self.detect_mode_from_filename(filename)
            
            # 資料清理
            df = df.dropna()
            df = df[df['Power'] >= 0]
            df['Mode'] = mode_name
            
            # 儲存到已載入檔案字典
            file_key = Path(file_path).name
            self.loaded_files[file_key] = {
                'path': file_path,
                'data': df,
                'mode': mode_name
            }
            
            # 更新檔案列表顯示
            self.update_file_list()
            
            # 更新下拉選單
            self.update_combo_boxes()
            
            self.update_status(f"成功載入: {file_key}")
            
        except Exception as e:
            messagebox.showerror("錯誤", f"載入檔案失敗: {str(e)}")
            self.update_status("載入失敗")
    
    def detect_mode_from_filename(self, filename):
        """從檔名推測模式"""
        
        filename_lower = filename.lower()
        
        if 'nolight' in filename_lower or 'no light' in filename_lower or '無燈' in filename_lower:
            return '無燈光'
        elif 'breath' in filename_lower or '呼吸' in filename_lower:
            return '呼吸燈'
        elif 'colorcycle' in filename_lower or 'color cycle' in filename_lower or 'color' in filename_lower or '彩色' in filename_lower:
            return '彩色循環'
        elif 'flash' in filename_lower or '閃爍' in filename_lower:
            return '閃爍'
        else:
            return '未知模式'
    
    def update_file_list(self):
        """更新檔案列表顯示"""
        
        # 清除現有項目
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
        
        # 添加已載入的檔案
        for file_key, file_info in self.loaded_files.items():
            df = file_info['data']
            
            # 計算統計資訊
            time_range = f"{df['Time'].min():.1f} - {df['Time'].max():.1f}s"
            avg_power = f"{df['Power'].mean()*1000:.2f} mW"
            
            self.file_tree.insert('', 'end', values=(
                file_key,
                file_info['mode'],
                len(df),
                time_range,
                avg_power
            ))
    
    def update_combo_boxes(self):
        """更新下拉選單和列表框"""
        
        file_names = list(self.loaded_files.keys())
        
        # 更新單檔分析下拉選單
        self.single_file_combo['values'] = file_names
        if file_names and not self.single_file_var.get():
            self.single_file_var.set(file_names[0])
        
        # 更新比較分析列表框
        self.comparison_listbox.delete(0, tk.END)
        for name in file_names:
            self.comparison_listbox.insert(tk.END, name)
    
    def clear_files(self):
        """清除所有已載入的檔案"""
        
        if messagebox.askyesno("確認", "確定要清除所有已載入的檔案嗎？"):
            self.loaded_files.clear()
            self.update_file_list()
            self.update_combo_boxes()
            self.update_status("已清除所有檔案")
    
    def remove_selected(self):
        """移除選中的檔案"""
        
        selected_items = self.file_tree.selection()
        if not selected_items:
            messagebox.showwarning("警告", "請先選擇要移除的檔案")
            return
        
        for item in selected_items:
            file_name = self.file_tree.item(item)['values'][0]
            if file_name in self.loaded_files:
                del self.loaded_files[file_name]
        
        self.update_file_list()
        self.update_combo_boxes()
        self.update_status(f"已移除 {len(selected_items)} 個檔案")
    
    def view_data(self):
        """檢視選中檔案的資料"""
        
        selected_items = self.file_tree.selection()
        if not selected_items:
            messagebox.showwarning("警告", "請先選擇要檢視的檔案")
            return
        
        item = selected_items[0]
        file_name = self.file_tree.item(item)['values'][0]
        
        if file_name in self.loaded_files:
            df = self.loaded_files[file_name]['data']
            
            # 建立新視窗顯示資料
            data_window = tk.Toplevel(self.root)
            data_window.title(f"資料檢視 - {file_name}")
            data_window.geometry("800x600")
            
            # 建立文字區域顯示資料摘要
            text_area = scrolledtext.ScrolledText(data_window)
            text_area.pack(fill='both', expand=True, padx=10, pady=10)
            
            # 顯示資料摘要
            summary = f"""檔案: {file_name}
模式: {self.loaded_files[file_name]['mode']}
資料點數: {len(df)}
時間範圍: {df['Time'].min():.2f} - {df['Time'].max():.2f} 秒

統計摘要:
平均電壓: {df['Voltage'].mean():.3f} V
平均電流: {df['Current'].mean()*1000:.2f} mA  
平均功率: {df['Power'].mean()*1000:.2f} mW
最大功率: {df['Power'].max()*1000:.2f} mW
最小功率: {df['Power'].min()*1000:.2f} mW

前10筆資料:
{df.head(10).to_string()}
"""
            
            text_area.insert('1.0', summary)
            text_area.config(state='disabled')

def main():
    """主函數"""
    
    # 建立主視窗
    root = tk.Tk()
    
    # 建立GUI應用程式
    app = MousePowerGUI(root)
    
    # 啟動GUI
    root.mainloop()

if __name__ == "__main__":
    main()   
 def run_single_analysis(self):
        """執行單檔分析"""
        
        selected_file = self.single_file_var.get()
        if not selected_file or selected_file not in self.loaded_files:
            messagebox.showwarning("警告", "請先選擇要分析的檔案")
            return
        
        # 在新執行緒中執行分析
        def analysis_thread():
            try:
                self.start_progress()
                self.update_status("執行單檔分析...")
                
                df = self.loaded_files[selected_file]['data']
                mode = self.loaded_files[selected_file]['mode']
                
                # 計算統計數據
                stats = self.calculate_statistics(df)
                
                # 生成分析圖表
                if self.include_detailed.get():
                    chart_file = self.create_single_analysis_chart(df, selected_file)
                
                # 準備結果文字
                result_text = self.format_single_analysis_result(stats, selected_file)
                
                # 在主執行緒中更新GUI
                self.root.after(0, lambda: self.display_single_result(result_text))
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("錯誤", f"分析失敗: {str(e)}"))
            finally:
                self.root.after(0, self.stop_progress)
                self.root.after(0, lambda: self.update_status("單檔分析完成"))
        
        threading.Thread(target=analysis_thread, daemon=True).start()
    
    def run_comparison_analysis(self):
        """執行多檔比較分析"""
        
        selected_indices = self.comparison_listbox.curselection()
        if len(selected_indices) < 2:
            messagebox.showwarning("警告", "請至少選擇2個檔案進行比較")
            return
        
        selected_files = [self.comparison_listbox.get(i) for i in selected_indices]
        
        def comparison_thread():
            try:
                self.start_progress()
                self.update_status("執行多檔比較分析...")
                
                # 準備資料
                data_dict = {}
                for file_name in selected_files:
                    if file_name in self.loaded_files:
                        data_dict[file_name] = self.loaded_files[file_name]['data']
                
                # 執行比較分析
                comparison_stats = self.calculate_comparison_statistics(data_dict)
                
                # 生成比較圖表
                if self.include_power_comparison.get():
                    chart_files = self.create_comparison_charts(data_dict)
                
                # 準備結果文字
                result_text = self.format_comparison_result(comparison_stats, selected_files)
                
                # 更新GUI
                self.root.after(0, lambda: self.display_comparison_result(result_text))
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("錯誤", f"比較分析失敗: {str(e)}"))
            finally:
                self.root.after(0, self.stop_progress)
                self.root.after(0, lambda: self.update_status("比較分析完成"))
        
        threading.Thread(target=comparison_thread, daemon=True).start()
    
    def calculate_statistics(self, df):
        """計算單檔統計數據"""
        
        duration = df['Time'].max() - df['Time'].min()
        total_energy = np.trapz(df['Power'], df['Time'])
        
        stats = {
            'duration_s': duration,
            'data_points': len(df),
            'avg_voltage_V': df['Voltage'].mean(),
            'avg_current_A': df['Current'].mean(),
            'avg_current_mA': df['Current'].mean() * 1000,
            'avg_power_W': df['Power'].mean(),
            'avg_power_mW': df['Power'].mean() * 1000,
            'max_power_W': df['Power'].max(),
            'max_power_mW': df['Power'].max() * 1000,
            'min_power_W': df['Power'].min(),
            'min_power_mW': df['Power'].min() * 1000,
            'std_power_W': df['Power'].std(),
            'std_power_mW': df['Power'].std() * 1000,
            'total_energy_J': total_energy,
            'cv_power': df['Power'].std() / df['Power'].mean() if df['Power'].mean() > 0 else 0
        }
        
        # 電池續航估算
        if self.include_battery.get():
            battery_capacity_mAh = 1000
            voltage = 3.7
            battery_energy_J = battery_capacity_mAh * voltage * 3.6
            
            if stats['avg_power_W'] > 0:
                stats['battery_hours'] = battery_energy_J / (stats['avg_power_W'] * 3600)
                stats['battery_days'] = stats['battery_hours'] / 24
            else:
                stats['battery_hours'] = float('inf')
                stats['battery_days'] = float('inf')
        
        return stats
    
    def calculate_comparison_statistics(self, data_dict):
        """計算多檔比較統計"""
        
        comparison_stats = {}
        
        # 計算每個檔案的統計數據
        for file_name, df in data_dict.items():
            comparison_stats[file_name] = self.calculate_statistics(df)
            comparison_stats[file_name]['mode'] = self.loaded_files[file_name]['mode']
        
        # 計算相對比較
        if len(data_dict) > 1:
            # 找出基準功率（通常是無燈光模式或最低功率）
            baseline_power = min(stats['avg_power_W'] for stats in comparison_stats.values())
            
            for file_name, stats in comparison_stats.items():
                if baseline_power > 0:
                    stats['power_increase_percent'] = ((stats['avg_power_W'] - baseline_power) / baseline_power * 100)
                else:
                    stats['power_increase_percent'] = 0
        
        return comparison_stats
    
    def create_single_analysis_chart(self, df, file_name):
        """建立單檔分析圖表"""
        
        # 設定中文字體
        plt.rcParams['font.family'] = ['Microsoft JhengHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
        
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle(f'{file_name} - 詳細分析', fontsize=16, fontweight='bold')
        
        # 功率時間序列
        axes[0,0].plot(df['Time'], df['Power']*1000, 'b-', linewidth=1, alpha=0.8)
        axes[0,0].set_title('功率變化時間序列')
        axes[0,0].set_xlabel('時間 (秒)')
        axes[0,0].set_ylabel('功率 (mW)')
        axes[0,0].grid(True, alpha=0.3)
        
        # 電流時間序列
        axes[0,1].plot(df['Time'], df['Current']*1000, 'g-', linewidth=1, alpha=0.8)
        axes[0,1].set_title('電流變化時間序列')
        axes[0,1].set_xlabel('時間 (秒)')
        axes[0,1].set_ylabel('電流 (mA)')
        axes[0,1].grid(True, alpha=0.3)
        
        # 功率分布直方圖
        axes[1,0].hist(df['Power']*1000, bins=30, alpha=0.7, edgecolor='black')
        axes[1,0].set_title('功率分布直方圖')
        axes[1,0].set_xlabel('功率 (mW)')
        axes[1,0].set_ylabel('頻次')
        axes[1,0].grid(True, alpha=0.3)
        
        # 累積能量消耗
        time_intervals = np.diff(np.concatenate([[0], df['Time']]))
        cumulative_energy = np.cumsum(df['Power'] * time_intervals)
        axes[1,1].plot(df['Time'], cumulative_energy, 'r-', linewidth=2)
        axes[1,1].set_title('累積能量消耗')
        axes[1,1].set_xlabel('時間 (秒)')
        axes[1,1].set_ylabel('累積能量 (J)')
        axes[1,1].grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        # 儲存圖表
        chart_file = f"{self.output_dir}/{Path(file_name).stem}_analysis.png"
        plt.savefig(chart_file, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_file
    
    def create_comparison_charts(self, data_dict):
        """建立比較分析圖表"""
        
        plt.rcParams['font.family'] = ['Microsoft JhengHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
        
        fig, axes = plt.subplots(2, 2, figsize=(16, 12))
        fig.suptitle('多檔案功率比較分析', fontsize=18, fontweight='bold')
        
        file_names = list(data_dict.keys())
        colors = plt.cm.Set3(np.linspace(0, 1, len(file_names)))
        
        # 平均功率比較
        avg_powers = [data_dict[name]['Power'].mean()*1000 for name in file_names]
        modes = [self.loaded_files[name]['mode'] for name in file_names]
        
        bars = axes[0,0].bar(modes, avg_powers, color=colors, alpha=0.8, edgecolor='black')
        axes[0,0].set_title('平均功率比較', fontweight='bold')
        axes[0,0].set_ylabel('平均功率 (mW)')
        axes[0,0].grid(True, alpha=0.3, axis='y')
        
        # 在柱狀圖上添加數值
        for bar, power in zip(bars, avg_powers):
            axes[0,0].text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1,
                          f'{power:.1f}', ha='center', va='bottom', fontweight='bold')
        
        # 功率分布箱型圖
        power_data = [data_dict[name]['Power']*1000 for name in file_names]
        bp = axes[0,1].boxplot(power_data, labels=modes, patch_artist=True)
        for patch, color in zip(bp['boxes'], colors):
            patch.set_facecolor(color)
            patch.set_alpha(0.7)
        axes[0,1].set_title('功率分布比較', fontweight='bold')
        axes[0,1].set_ylabel('功率 (mW)')
        axes[0,1].grid(True, alpha=0.3, axis='y')
        
        # 時間序列疊加圖
        for i, (name, df) in enumerate(data_dict.items()):
            mode = self.loaded_files[name]['mode']
            # 正規化時間到0-100%
            time_norm = (df['Time'] - df['Time'].min()) / (df['Time'].max() - df['Time'].min()) * 100
            axes[1,0].plot(time_norm, df['Power']*1000, 
                          color=colors[i], label=mode, alpha=0.8, linewidth=2)
        
        axes[1,0].set_title('功率時間序列比較', fontweight='bold')
        axes[1,0].set_xlabel('時間進度 (%)')
        axes[1,0].set_ylabel('功率 (mW)')
        axes[1,0].legend()
        axes[1,0].grid(True, alpha=0.3)
        
        # 電池續航比較
        if self.include_battery_comparison.get():
            battery_capacity_mAh = 1000
            voltage = 3.7
            battery_energy_J = battery_capacity_mAh * voltage * 3.6
            
            battery_hours = []
            for name in file_names:
                avg_power_W = data_dict[name]['Power'].mean()
                hours = battery_energy_J / (avg_power_W * 3600) if avg_power_W > 0 else 0
                battery_hours.append(hours)
            
            bars = axes[1,1].bar(modes, battery_hours, color=colors, alpha=0.8, edgecolor='black')
            axes[1,1].set_title(f'預估電池續航 ({battery_capacity_mAh}mAh)', fontweight='bold')
            axes[1,1].set_ylabel('續航時間 (小時)')
            axes[1,1].grid(True, alpha=0.3, axis='y')
            
            for bar, hours in zip(bars, battery_hours):
                axes[1,1].text(bar.get_x() + bar.get_width()/2, bar.get_height() + 5,
                              f'{hours:.0f}h', ha='center', va='bottom', fontweight='bold')
        
        plt.tight_layout()
        
        # 儲存圖表
        chart_file = f"{self.output_dir}/comparison_analysis.png"
        plt.savefig(chart_file, dpi=300, bbox_inches='tight')
        plt.close()
        
        return [chart_file]
    
    def format_single_analysis_result(self, stats, file_name):
        """格式化單檔分析結果"""
        
        mode = self.loaded_files[file_name]['mode']
        
        result = f"""=== {file_name} 分析結果 ===

檔案資訊:
• 模式: {mode}
• 測量時間: {stats['duration_s']:.1f} 秒
• 資料點數: {stats['data_points']} 筆

電氣參數:
• 平均電壓: {stats['avg_voltage_V']:.3f} V
• 平均電流: {stats['avg_current_mA']:.2f} mA
• 平均功率: {stats['avg_power_mW']:.2f} mW
• 最大功率: {stats['max_power_mW']:.2f} mW
• 最小功率: {stats['min_power_mW']:.2f} mW
• 功率標準差: {stats['std_power_mW']:.2f} mW

能量分析:
• 總消耗能量: {stats['total_energy_J']:.3f} J
• 功率變異係數: {stats['cv_power']:.3f}
"""

        if self.include_battery.get() and 'battery_hours' in stats:
            result += f"""
電池續航估算 (1000mAh, 3.7V):
• 預估續航: {stats['battery_hours']:.1f} 小時
• 預估續航: {stats['battery_days']:.1f} 天
"""

        result += f"""
分析完成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
        
        return result
    
    def format_comparison_result(self, comparison_stats, selected_files):
        """格式化比較分析結果"""
        
        result = f"""=== 多檔案比較分析結果 ===

比較檔案數量: {len(selected_files)}

"""
        
        # 各檔案統計摘要
        for file_name in selected_files:
            if file_name in comparison_stats:
                stats = comparison_stats[file_name]
                result += f"""【{stats['mode']}】 - {file_name}
• 平均功率: {stats['avg_power_mW']:.2f} mW
• 平均電流: {stats['avg_current_mA']:.2f} mA
• 測量時間: {stats['duration_s']:.1f} 秒
• 相對增加: {stats.get('power_increase_percent', 0):+.1f}%
"""
                if 'battery_hours' in stats:
                    result += f"• 預估續航: {stats['battery_hours']:.1f} 小時\n"
                result += "\n"
        
        # 比較摘要
        if len(comparison_stats) > 1:
            powers = [stats['avg_power_mW'] for stats in comparison_stats.values()]
            modes = [stats['mode'] for stats in comparison_stats.values()]
            
            max_power_idx = powers.index(max(powers))
            min_power_idx = powers.index(min(powers))
            
            result += f"""比較摘要:
• 最高功耗模式: {modes[max_power_idx]} ({powers[max_power_idx]:.2f} mW)
• 最低功耗模式: {modes[min_power_idx]} ({powers[min_power_idx]:.2f} mW)
• 功耗差異: {max(powers) - min(powers):.2f} mW ({((max(powers) - min(powers))/min(powers)*100):.1f}%)

分析完成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
        
        return result
    
    def display_single_result(self, result_text):
        """顯示單檔分析結果"""
        
        self.single_result_text.delete('1.0', tk.END)
        self.single_result_text.insert('1.0', result_text)
    
    def display_comparison_result(self, result_text):
        """顯示比較分析結果"""
        
        self.comp_result_text.delete('1.0', tk.END)
        self.comp_result_text.insert('1.0', result_text)
    
    def select_all_files(self):
        """選擇所有檔案"""
        
        self.comparison_listbox.select_set(0, tk.END)
    
    def clear_selection(self):
        """清除選擇"""
        
        self.comparison_listbox.selection_clear(0, tk.END)
    
    def select_output_dir(self):
        """選擇輸出目錄"""
        
        directory = filedialog.askdirectory(title="選擇輸出目錄")
        if directory:
            self.output_dir = directory
            self.output_dir_var.set(directory)
    
    def generate_full_report(self):
        """生成完整報告"""
        
        if not self.loaded_files:
            messagebox.showwarning("警告", "請先載入檔案")
            return
        
        def report_thread():
            try:
                self.start_progress()
                self.update_status("生成完整報告...")
                
                # 更新輸出目錄
                self.output_dir = self.output_dir_var.get()
                Path(self.output_dir).mkdir(exist_ok=True)
                
                report_content = self.create_full_report()
                
                # 儲存報告
                if self.export_excel.get():
                    self.save_excel_report()
                
                # 更新報告預覽
                self.root.after(0, lambda: self.display_report_preview(report_content))
                
                self.root.after(0, lambda: messagebox.showinfo("完成", f"報告已生成至: {self.output_dir}"))
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("錯誤", f"報告生成失敗: {str(e)}"))
            finally:
                self.root.after(0, self.stop_progress)
                self.root.after(0, lambda: self.update_status("報告生成完成"))
        
        threading.Thread(target=report_thread, daemon=True).start()
    
    def create_full_report(self):
        """建立完整報告內容"""
        
        report = f"""無線滑鼠耗電分析報告
生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
分析檔案數量: {len(self.loaded_files)}

=== 檔案清單 ===
"""
        
        for file_name, file_info in self.loaded_files.items():
            df = file_info['data']
            stats = self.calculate_statistics(df)
            
            report += f"""
檔案: {file_name}
模式: {file_info['mode']}
平均功率: {stats['avg_power_mW']:.2f} mW
測量時間: {stats['duration_s']:.1f} 秒
資料點數: {stats['data_points']} 筆
"""
        
        # 如果有多個檔案，進行比較分析
        if len(self.loaded_files) > 1:
            data_dict = {name: info['data'] for name, info in self.loaded_files.items()}
            comparison_stats = self.calculate_comparison_statistics(data_dict)
            
            report += "\n=== 比較分析 ===\n"
            
            powers = [stats['avg_power_mW'] for stats in comparison_stats.values()]
            modes = [stats['mode'] for stats in comparison_stats.values()]
            
            if powers:
                max_power_idx = powers.index(max(powers))
                min_power_idx = powers.index(min(powers))
                
                report += f"""
最高功耗: {modes[max_power_idx]} ({powers[max_power_idx]:.2f} mW)
最低功耗: {modes[min_power_idx]} ({powers[min_power_idx]:.2f} mW)
功耗範圍: {max(powers) - min(powers):.2f} mW
"""
        
        return report
    
    def save_excel_report(self):
        """儲存Excel報告"""
        
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "耗電分析報告"
            
            # 標題行
            headers = ['檔案名稱', '模式', '平均功率(mW)', '最大功率(mW)', '平均電流(mA)', 
                      '測量時間(s)', '資料點數', '總能量(J)', '預估續航(h)']
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 資料行
            for row, (file_name, file_info) in enumerate(self.loaded_files.items(), 2):
                df = file_info['data']
                stats = self.calculate_statistics(df)
                
                ws.cell(row=row, column=1, value=file_name)
                ws.cell(row=row, column=2, value=file_info['mode'])
                ws.cell(row=row, column=3, value=round(stats['avg_power_mW'], 2))
                ws.cell(row=row, column=4, value=round(stats['max_power_mW'], 2))
                ws.cell(row=row, column=5, value=round(stats['avg_current_mA'], 2))
                ws.cell(row=row, column=6, value=round(stats['duration_s'], 1))
                ws.cell(row=row, column=7, value=stats['data_points'])
                ws.cell(row=row, column=8, value=round(stats['total_energy_J'], 3))
                
                if 'battery_hours' in stats:
                    ws.cell(row=row, column=9, value=round(stats['battery_hours'], 1))
            
            # 儲存檔案
            excel_file = f"{self.output_dir}/耗電分析報告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(excel_file)
            
        except ImportError:
            messagebox.showwarning("警告", "需要安裝 openpyxl 套件才能匯出Excel報告")
        except Exception as e:
            messagebox.showerror("錯誤", f"Excel報告儲存失敗: {str(e)}")
    
    def display_report_preview(self, report_content):
        """顯示報告預覽"""
        
        self.report_text.delete('1.0', tk.END)
        self.report_text.insert('1.0', report_content)
    
    def open_output_dir(self):
        """開啟輸出目錄"""
        
        output_path = self.output_dir_var.get()
        if os.path.exists(output_path):
            if sys.platform == "win32":
                os.startfile(output_path)
            elif sys.platform == "darwin":
                os.system(f"open '{output_path}'")
            else:
                os.system(f"xdg-open '{output_path}'")
        else:
            messagebox.showwarning("警告", "輸出目錄不存在")